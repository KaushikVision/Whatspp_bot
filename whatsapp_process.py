from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.edge.options import Options
from datetime import datetime
from openpyxl.styles import NamedStyle
import time
import os
import xlwings as xw
import openpyxl
import logging
from constant import Constant
import pythoncom
import glob
from dateutil.relativedelta import relativedelta
import wmi
from openpyxl.styles import PatternFill
import win32com.client
import re 
import traceback
class WhatsappProcess:
    running_whatsapp = True
    Whatsapp_excel_file = Constant.Whatsapp_excel_file

    def kill_excel(self):
        pythoncom.CoInitialize()
        w = wmi.WMI()
        mainWindowStatus = False
        processName = 'excel.exe *32'
        for process in w.Win32_process():
            if(processName == str(process.Name)):
                mainWindowStatus = True
                break
        if(mainWindowStatus):
            os.system(f'taskkill /f /im excel.exe *32')

    def IS_LOCKED(self, filepath):
        try:
            locked = None
            file_object = None
            if os.path.exists(filepath):
                try:
                    buffer_size = 8
                    file_object = open(filepath, 'a', buffer_size)
                    if file_object:
                        locked = False
                except IOError:
                    logging.info(f"{filepath} File is locked (unable to open in append mode)")
                    locked = True
                finally:
                    if file_object:
                        file_object.close()
            else:
                logging.info(f"File Path Not Found {filepath}")

            return locked
        except Exception as error:
            logging.info("Function: IS_LOCKED (Function IS_LOCKED Error)")

    def WAIT_FOR_FILE(self, filepaths):
        try:
            wait_time = 5
            for filepath in filepaths:
                while not os.path.exists(filepath):
                    time.sleep(wait_time)
                while self.IS_LOCKED(filepath):
                    time.sleep(wait_time)
        except Exception as error:
            logging.info(f'WAIT_FOR_FILE : {error}')
            pass

    def Send_data(self, Workbook, total_row, sheet, driver, date_string, excel_name):
        print('send message on whatsapp')
        valid_time = False
        try:
            try:
                if self.running_whatsapp:
                    for row_data in range(2, total_row):
                        col_a = sheet.Cells(row_data, 1).Value
                        if(str(col_a) == '1.0'):
                            print('\n', '1111', '\n')
                            wbExceptTemp = openpyxl.load_workbook(excel_name)
                            wbSheet = wbExceptTemp['MESSAGES']
                            col_b = wbSheet.cell(row_data, 2).value
                            col_g = wbSheet.cell(row_data, 7).value
                            col_h = wbSheet.cell(row_data, 8).value
                            wbExceptTemp.close()
                            print('2222')
                            current_time = datetime.now()
                            current_time_str = current_time.strftime("%d/%m/%Y %H:%M:%S")
                            current_time_date = datetime.strptime(current_time_str, "%d/%m/%Y %H:%M:%S")
                            col_b = str(col_b).strip().upper()
                            
                            print('3333')
                            if(col_b == 'NONE' or col_b == ''):
                                if((str(col_g) == 'None' or str(col_g) == '') and (str(col_h) == 'None' or str(col_h) == '')):
                                    valid_time = True
                                else:
                                    if(str(col_g) <= str(current_time_date.time()) <= str(col_h)):
                                        valid_time = True
                                        
                            print('4444')
                            if((col_b != 'NONE' and col_b != '')):
                                col_b = datetime.fromisoformat(col_b)
                                col_b = col_b.strftime("%d/%m/%Y %H:%M:%S")
                                col_b_date = datetime.strptime(col_b, "%d/%m/%Y %H:%M:%S")
                                if(col_b_date <= current_time_date):
                                    if((str(col_g) == 'None' or str(col_g) == '') and (str(col_h) == 'None' or str(col_h) == '')):
                                        valid_time = True
                                    else:
                                        if(str(col_g) <= str(current_time_date.time()) < str(col_h)):
                                            valid_time = True
                            
                            print('5555', valid_time)
                            if(valid_time):
                                groupName = sheet.Cells(row_data, 16).Value
                                sendText = sheet.Cells(row_data, 19).Value
                                if(str(type(sendText)) == "<class 'pywintypes.datetime'>"):
                                    sendText = str(sendText).split('.')[0]
                                
                                print('6666')
                                sendText = str(sendText)
                                
                                print(f'sendText : {sendText}')
                                searchInput = WebDriverWait(driver, 300).until(EC.presence_of_element_located(
                                    (By.XPATH, '''//div[@id='side']//div[1]//p''')))
                                print('7777')
                                groupName = re.sub(r'[\U0001F600-\U0001F64F\U0001F300-\U0001F5FF\U0001F680-\U0001F6FF\U0001F700-\U0001F77F\U0001F780-\U0001F7FF\U0001F800-\U0001F8FF\U0001F900-\U0001F9FF\U0001FA00-\U0001FA6F\U0001FA70-\U0001FAFF\U0001FB00-\U0001FBFF\U0001FC00-\U0001FCFF\U0001FD00-\U0001FDFF\U0001FE00-\U0001FEFF\U0001FF00-\U0001FFFF\u2600-\u26FF\u2700-\u27BF]+', '', groupName)
                                print('8888', groupName)
                                groupName = str(groupName).strip()
                                searchInput.send_keys(Keys.CONTROL, 'a')
                                print('9999')
                                searchInput.send_keys(groupName)
                                print('10101010')
                                time.sleep(3)
                                print('&&&&&')
                                # find_proper = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located(
                                    # (By.XPATH, f'''//span[contains(@title, '{groupName}')''')))
                                # print(find_proper, '$$$$')
                                senMessageStatus = False
                                
                                asci_symbol = ''.join(chr(i) for i in range(32,127) if not chr(i).isalnum())
                                if groupName[-1] in asci_symbol:
                                    groupName = groupName[0:(len(groupName)-1)]
                                   
                                logging.info(f'group name {groupName}')   
                                print('****************************')
                                print('****************************')
                                print(f"//span[contains(@title , '{groupName}')]/span[contains(text(),'{groupName}')]")
                                try:
                                    """select the contact"""
                                    WebDriverWait(driver, 10).until(
                                        EC.presence_of_element_located(
                                            (
                                                By.XPATH,
                                                f"//span[contains(@title , '{groupName}')]/span[contains(text(),'{groupName}')]",
                                            )
                                        )
                                    ).click()
                                    senMessageStatus = True
                                    time.sleep(2)
                                    print('select the contact')
                                except Exception as e:  # noqa
                                    logging.info(f'error in click on groupName : {e}')
                                    pass
                                
                                
                                # for group_name in find_proper:
                                    # grp_name = group_name.get_attribute('title')
                                    # grp_name =  re.sub(r'[\U0001F600-\U0001F64F\U0001F300-\U0001F5FF\U0001F680-\U0001F6FF\U0001F700-\U0001F77F\U0001F780-\U0001F7FF\U0001F800-\U0001F8FF\U0001F900-\U0001F9FF\U0001FA00-\U0001FA6F\U0001FA70-\U0001FAFF\U0001FB00-\U0001FBFF\U0001FC00-\U0001FCFF\U0001FD00-\U0001FDFF\U0001FE00-\U0001FEFF\U0001FF00-\U0001FFFF\u2600-\u26FF\u2700-\u27BF]+', '', grp_name)
                                    # grp_name = str(grp_name).strip()
                                    # if(grp_name and (grp_name).strip() == str(groupName).strip()):
                                        # group_name.click()
                                        # senMessageStatus = True
                                        # time.sleep(2)
                                        # break

                                if(senMessageStatus):
                                    colI_year = sheet.Cells(row_data, 9).Value
                                    colJ_year = sheet.Cells(row_data, 10).Value
                                    colK_year = sheet.Cells(row_data, 11).Value
                                    colL_year = sheet.Cells(row_data, 12).Value
                                    colM_year = sheet.Cells(row_data, 13).Value
                                    colN_year = sheet.Cells(row_data, 14).Value
                                    if((str(colI_year) == '' or (str(colI_year) == 'None'))):
                                        fill_cell = PatternFill(patternType='solid', fgColor=Constant.Red_color_code)
                                        sheet.Cells(row_data, 9).fill = fill_cell
                                        continue
                                    if((str(colJ_year) == '' or (str(colJ_year) == 'None'))):
                                        fill_cell = PatternFill(patternType='solid', fgColor=Constant.Red_color_code)
                                        sheet.Cells(row_data, 10).fill = fill_cell
                                        continue
                                    if((str(colK_year) == '' or (str(colK_year) == 'None'))):
                                        fill_cell = PatternFill(patternType='solid', fgColor=Constant.Red_color_code)
                                        sheet.Cells(row_data, 11).fill = fill_cell
                                        continue
                                    if((str(colL_year) == '' or (str(colL_year) == 'None'))):
                                        fill_cell = PatternFill(patternType='solid', fgColor=Constant.Red_color_code)
                                        sheet.Cells(row_data, 12).fill = fill_cell
                                        continue
                                    if((str(colM_year) == '' or (str(colM_year) == 'None'))):
                                        fill_cell = PatternFill(patternType='solid', fgColor=Constant.Red_color_code)
                                        sheet.Cells(row_data, 13).fill = fill_cell
                                        continue
                                    if((str(colN_year) == '' or (str(colN_year) == 'None'))):
                                        fill_cell = PatternFill(patternType='solid', fgColor=Constant.Red_color_code)
                                        sheet.Cells(row_data, 14).fill = fill_cell
                                        continue

                                    # sendInput = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                        # (By.XPATH, "//div[@role='textbox' and @aria-label='Type a message']")))
                                    sendInput = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@role='textbox' and @aria-placeholder='Type a message']")))

                                    fileType = sheet.Cells(row_data, 17).Value 
                                    messageText = sheet.Cells(row_data, 19).Value 
                                    if(str(fileType) == 'None'):
                                        send_status = False
                                        searchInput.send_keys(Keys.CONTROL, 'a')
                                        try:
                                            for one_line in sendText.split("\n"):
                                                sendInput.send_keys(one_line)
                                                sendInput.send_keys(Keys.SHIFT + Keys.ENTER)
                                            sendInput.send_keys(Keys.ENTER)
                                            send_status = True
                                        except:
                                            sendInput.send_keys(Keys.CONTROL, 'a')
                                            sendInput.send_keys(Keys.BACK_SPACE)
                                            send_status = False

                                        if(not send_status):
                                            driver.execute_script(
                                                    f'''
                                                const text = `{sendText}`;
                                                const dataTransfer = new DataTransfer();
                                                dataTransfer.setData('text', text);
                                                const event = new ClipboardEvent('paste', {{
                                                clipboardData: dataTransfer,
                                                bubbles: true
                                                }});
                                                arguments[0].dispatchEvent(event)
                                                ''',
                                                    sendInput)
                                            sendInput.send_keys(Keys.ENTER)

                                    elif(str(fileType).strip() == 'Photos and Videos' and (str(messageText).strip() != 'None' and str(messageText).strip() != '')):
                                        sendFilePath = sheet.Cells(row_data, 18).Value
                                        attachBtn = WebDriverWait(driver, 5).until(EC.presence_of_element_located(
                                            (By.XPATH, "//div[contains(@title, 'Attach')]")))
                                        attachBtn.click()
                                        fileSize = os.path.getsize(sendFilePath)
                                        if(sendFilePath.__contains__("png") or sendFilePath.__contains__("jpeg") or sendFilePath.__contains__("jpg")):
                                            if(fileSize <= 1000000):
                                                multiImg = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                                        (By.XPATH, "//input[@type='file'][@accept='image/*,video/mp4,video/3gpp,video/quicktime']")))
                                            else:
                                                multiImg = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                                        (By.XPATH, "//input[@type='file'][@accept='*']")))
                                        else:
                                            multiImg = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                (By.XPATH, "//input[@type='file'][@accept='*']")))

                                        multiImg.send_keys(sendFilePath)
                                        time.sleep(2)

                                        sendInput = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located(
                                        (By.XPATH, "//div[contains(@data-testid, 'media-caption-input-container')]")))
                                        sendInput[0].send_keys(messageText)
                                        sendingBtn = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                            (By.XPATH, "//span[contains(@data-icon, 'send')]")))
                                        sendingBtn.click()
                                        time.sleep(2)
                                    elif(str(fileType).strip() == 'Photos and Videos' and (str(messageText).strip() == 'None' or str(messageText).strip() == '')):
                                        sendFilePath = sheet.Cells(row_data, 18).Value
                                        # import pdb;pdb.set_trace()
                                        # attachBtn = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//div[contains(@title, 'Attach')]")))
                                        attachBtn = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//button[contains(@title, 'Attach')]")))
                                        attachBtn.click()
                                        fileSize = os.path.getsize(sendFilePath)
                                        if(sendFilePath.__contains__("png") or sendFilePath.__contains__("jpeg") or sendFilePath.__contains__("jpg")):
                                            if(fileSize <= 1000000):
                                                multiImg = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                                        (By.XPATH, "//input[@type='file'][@accept='image/*,video/mp4,video/3gpp,video/quicktime']")))
                                            else:
                                                multiImg = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                                        (By.XPATH, "//input[@type='file'][@accept='*']")))
                                        else:
                                            multiImg = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                (By.XPATH, "//input[@type='file'][@accept='*']")))

                                        multiImg.send_keys(sendFilePath)
                                        time.sleep(2)
                                        sendingBtn = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                            (By.XPATH, "//span[contains(@data-icon, 'send')]")))
                                        sendingBtn.click()
                                        time.sleep(2)
                                    else:
                                        searchInput.send_keys(Keys.CONTROL, 'a')
                                        sendInput.send_keys(sendText)
                                        sendInput.send_keys(Keys.RETURN)

                                    duration = 'msg-time'
                                    error_count = 0
                                    try:
                                        while(duration == "msg-time" and error_count != 3):
                                            time.sleep(2)
                                            try:
                                                pp = WebDriverWait(driver, 180).until(EC.presence_of_all_elements_located(
                                                    (By.XPATH, "//span[contains(@data-icon, 'msg-')]")))
                                                # print('pp', pp)
                                                duration = pp[-1].get_attribute(
                                                    "data-icon")
                                            except:
                                                time.sleep(10)
                                                error_count += 1

                                    except Exception as error:
                                        logging.info(
                                            "Function: Send_data (Message sending time error Please check Internet Connection)")
                                        logging.error(error)

                                    if(duration != "msg-time" or  error_count < 3):
                                        now_datetime = datetime.now()
                                        date_string = now_datetime.strftime("%d/%m/%Y %H:%M:%S")
                                        current_time_date = datetime.strptime(date_string, "%d/%m/%Y %H:%M:%S")
                                        # date_style = Workbook.Styles.Add("ddmmyy")
                                        # date_style.NumberFormat = "DDMMYY"
                                        sheet.Cells(row_data, 20).NumberFormat = 'DDMMYY'
                                        sheet.Cells(row_data, 20).Value = date_string
                                        colI_year = sheet.Cells(row_data, 9).Value
                                        colJ_year = sheet.Cells(row_data, 10).Value
                                        colK_year = sheet.Cells(row_data, 11).Value
                                        colL_year = sheet.Cells(row_data, 12).Value
                                        colM_year = sheet.Cells(row_data, 13).Value
                                        colN_year = sheet.Cells(row_data, 14).Value
                                        
                                        future_date = current_time_date + relativedelta(years=int(colI_year), months=int(colJ_year), days=int(colK_year), hours=int(colL_year), minutes=int(colM_year), seconds=int(colN_year))
                                        
                                        sheet.Cells(row_data, 2).Value = str(future_date)
                                        if(str(Constant.Whatsapp_final_path) in str(excel_name)):
                                            sheet.Cells(row_data, 1).Value = 0

            except Exception as error:
                import traceback
                print('Send_data sub Error :-', error)
                logging.info(f'Send_data sub Error :- {traceback.format_exc()}')
                try:
                    if(self.running_whatsapp and valid_time):
                        time.sleep(2)
                        checkFiveTime = 0
                        while checkFiveTime != 5:
                            time.sleep(5)
                            logout_retry = WebDriverWait(driver, 20).until(
                                EC.presence_of_all_elements_located((By.XPATH, "//div[@role='button']//div//div")))
                            retry_text = logout_retry[1].text
                            if(str(retry_text).upper() == 'RETRY NOW'):
                                driver.execute_script(
                                    "arguments[0].click()", logout_retry[1])

                            checkFiveTime += 1

                        if(checkFiveTime == 5):
                            logging.info("Internet Connection Lost")

                except Exception as error:
                    print('Send_data 1) Error :- ', error)
                    try:
                        pythoncom.CoInitialize()
                        excel_app = xw.App(visible=False)
                        excel_book = excel_app.books.open(
                            excel_name)
                        excel_book.save()
                        excel_book.close()
                        excel_app.quit()
                        searchInput = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                            (By.XPATH, "//div[contains(@role, 'textbox')][contains(@title, 'Search input textbox')]")))
                        driver.close()
                        self.Send_data(Workbook, total_row,
                                       sheet, driver, date_string, excel_name)
                    except Exception as error:
                        print('Send_data 2) Error :- ', error)
                        # wbExceptTemp = openpyxl.load_workbook(
                        #     excel_name)
                        # self.WAIT_FOR_FILE(
                        #     [excel_name])
                        # wbExceptTemp.save(excel_name)
                        driver.close()
                        logging.info("Whatsapp Connection Lost error")
                        logging.info(f"error : {traceback.format_exc()}")

        except Exception as error:
            print('Send_data Main Error :- ', error)
            logging.info("Function: Send_data (Whatsapp Data sending time error)")
            # time.sleep(60)
            # self.kill_excel()

    def get_total_row(self, wbsheet):
        total_row = 0
        for i in range(2, wbsheet.Rows.Count+1):
            data = wbsheet.Cells(i, 1).Value
            if(str(data) == 'None' or str(data) == ''):
                total_row = i
                break
        if(total_row == 0):
            total_row = wbsheet.Rows.Count

        return total_row

    def check_all_sended_or_not(self, total_row, wbsheet, excel_name=None):
        all_cola_val = []
        if(wbsheet):
            sheet_name = wbsheet
            for row_data in range(2, total_row):
                col_a = wbsheet.Cells(row_data, 1).Value
                all_cola_val.append(str(col_a))

        return all_cola_val

    def close_edge_instance(self):
         # close edge instance
        w = wmi.WMI()
        mainWindowStatus = False
        processName = 'msedge.exe'
        for process in w.Win32_process():
            if(processName == str(process.Name)):
                mainWindowStatus = True
                break
        if(mainWindowStatus):
            os.system(f'taskkill /f /im msedge.exe')
            time.sleep(2)
        
    def whatsapp_auto_04(self):
        try:
            print('whatsapp_auto_04')
            Constant.Current_running_process = 'Whatsapp'
            if(WhatsappProcess.running_whatsapp):   
                if(Constant.Whatsapp_excel_file != ''):
                    pythoncom.CoInitialize()
                    self.close_edge_instance()
                    self.WAIT_FOR_FILE([Constant.Whatsapp_excel_file])
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    wbdata = excel.Workbooks.Open(Constant.Whatsapp_excel_file)
                    wbsheet = wbdata.Sheets["MESSAGES"]

                    # wbdata = openpyxl.load_workbook(Constant.Whatsapp_excel_file, data_only=True)
                    # wbsheet = wbdata['MESSAGES']

                    total_row = self.get_total_row(wbsheet)
                    print('*******1********')
                    all_cola_val = self.check_all_sended_or_not(total_row, wbsheet)
                    if(not all(str(v) == '0.0' for v in all_cola_val)):
                        # reload excel
                        pythoncom.CoInitialize()
                        excel_app = xw.App(visible=False)
                        excel_book = excel_app.books.open(Constant.Whatsapp_excel_file)
                        excel_book.save()
                        excel_book.close()
                        excel_app.quit()
                    else:
                        time.sleep(5)
                    print('*******2********')
                    now = datetime.now()
                    date_string = now.strftime("%d/%m/%Y %H:%M:%S")
                    try:
                        print('*******3********')
                        if(not all(v == '0.0' for v in all_cola_val)):
                            print('*******4********')
                            edge_options = Options()
                            edge_options.add_argument(
                                "--user-data-dir=" + 'C:\\Users\\Mepani Desktop 3\\AppData\\Local\\Microsoft\\Edge\\User Data\\Profile 1')
                            edge_options.add_experimental_option(
                                "excludeSwitches", ["enable-automation"])
                            edge_options.add_experimental_option(
                                'useAutomationExtension', False)
                            edge_options.add_argument(
                                '--disable-blink-features=AutomationControlled')

                            self.driver = webdriver.Edge(options=edge_options)
                            self.driver.maximize_window()
                            self.driver.get("https://web.whatsapp.com/")
                            print('\n', '$$$$$$$$$$$$$$$$$$$$$$$$$$$', '\n')
                            self.Send_data(wbdata, total_row, wbsheet,
                                   self.driver, date_string, Constant.Whatsapp_excel_file)
                            time.sleep(5)
                            wbdata.Save()
                            wbdata.Close(False)
                            excel.Quit()
                            self.WAIT_FOR_FILE([Constant.Whatsapp_excel_file])
                            logging.info('************ BEFORE ******************')
                            self.driver.close()
                            logging.info('************ AFTER ******************')
                        else:
                            print('*******5********')
                            # wbdata.Save()
                            wbdata.Close(False)
                            excel.Quit()
                            self.WAIT_FOR_FILE([Constant.Whatsapp_excel_file])
                    except Exception as error:
                        print('whatsapp_auto_04 Error 1:- ', error)
                        logging.info('whatsapp_auto_04 Error 1:- ', error)
                        time.sleep(5)
                        wbdata.Save()
                        wbdata.Close(False)
                        excel.Quit()
                        self.WAIT_FOR_FILE([Constant.Whatsapp_excel_file])


                if(Constant.Whatsapp_final_path != ''):
                    listFile = glob.glob(Constant.Whatsapp_final_path + './*.xlsx', recursive=True)
                
                if(len(listFile) != 0):
                    for findOne in listFile:
                        pythoncom.CoInitialize()
                        self.close_edge_instance()
                        new_excel = str(findOne).split('\\')[-1]
                        self.WAIT_FOR_FILE([findOne])
                        excel = win32com.client.Dispatch("Excel.Application")
                        excel.Visible = False
                        wbdata = excel.Workbooks.Open(findOne)
                        wbsheet = wbdata.Sheets["MESSAGES"]


                        # wbdata = openpyxl.load_workbook(findOne, data_only=True)
                        # wbsheet = wbdata['MESSAGES']

                        total_row = self.get_total_row(wbsheet)
                        all_cola_val = self.check_all_sended_or_not(total_row, wbsheet)
                        if(not all(str(v) == '0.0' for v in all_cola_val)):
                            # reload excel
                            pythoncom.CoInitialize()
                            new_excel_app = xw.App(visible=False)
                            new_excel_book = new_excel_app.books.open(findOne)
                            new_excel_book.save()
                            new_excel_book.close()
                            new_excel_app.quit()
                        else:
                            time.sleep(5)

                        now = datetime.now()
                        date_string = now.strftime("%d/%m/%Y %H:%M:%S")
                        try:
                            if(not all(v == '0.0' for v in all_cola_val)):
                                edge_options = Options()
                                edge_options.add_argument(
                                    "--user-data-dir=" + 'C:\\Users\\Mepani Desktop 3\\AppData\\Local\\Microsoft\\Edge\\User Data\\Profile 1')
                                edge_options.add_experimental_option(
                                    "excludeSwitches", ["enable-automation"])
                                edge_options.add_experimental_option(
                                    'useAutomationExtension', False)
                                edge_options.add_argument(
                                    '--disable-blink-features=AutomationControlled')

                                self.driver = webdriver.Edge(options=edge_options)
                                self.driver.maximize_window()
                                self.driver.get("https://web.whatsapp.com/")
                                print('\n', '&&&&&&&&&&&&&&&&&&&&&&&&&&&&&', '\n')
                                self.Send_data(wbdata, total_row, wbsheet,
                                    self.driver, date_string, findOne)
                                time.sleep(5)
                                self.driver.close()

                                # check all message sended or not
                                all_cola_val = self.check_all_sended_or_not(total_row, None, findOne)
                                if(all(str(v) == '0.0' for v in all_cola_val)):
                                    ss_folder_path = f'{Constant.Whatsapp_final_path}\\SS'
                                    if(not os.path.exists(ss_folder_path)):
                                        os.mkdir(ss_folder_path)

                                    wbdata.Save()
                                    wbdata.Close(False)
                                    excel.Quit()
                                    self.WAIT_FOR_FILE([Constant.Whatsapp_excel_file])
                                    rename_new_file = str(new_excel).split('.')[0]

                                    newDate = datetime.now()
                                    if(str(newDate).__contains__('.')):
                                        newDate = str(newDate).replace(':', '-')
                                        newDate = str(newDate).split('.')[0]
                                    else:
                                        newDate = str(newDate)

                                    os.rename(findOne, Constant.Whatsapp_final_path + f'\\SS\\{rename_new_file} {newDate}.xlsx')
                                else:
                                    # wbdata.Save()
                                    wbdata.Close(False)
                                    excel.Quit()
                                    self.WAIT_FOR_FILE([Constant.Whatsapp_excel_file])
                        except Exception as error:
                            print('whatsapp_auto_04 Error 2:- ', error)
                            time.sleep(5)
                            
                        

                
        except Exception as error:
            print('whatsapp_auto_04 Error 3:- ', error)
            
            logging.info(f"Function: whatsapp_auto_04 : {traceback.format_exc()}")
